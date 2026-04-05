"""
Bulk Operations Utilities
Import/Export data in CSV and Excel formats
"""
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.db import transaction
from datetime import datetime


class BulkImporter:
    """
    Handle bulk import operations
    """

    @staticmethod
    def import_users_csv(csv_file, role='STUDENT'):
        """
        Import users from CSV file
        Expected columns: email, first_name, last_name, phone, date_of_birth, gender
        """
        from .models import User

        results = {
            'success': [],
            'errors': [],
            'total': 0
        }

        try:
            # Read CSV
            decoded_file = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(decoded_file))

            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):
                    results['total'] += 1

                    try:
                        # Validate required fields
                        if not row.get('email'):
                            results['errors'].append({
                                'row': row_num,
                                'error': 'Email is required'
                            })
                            continue

                        # Check if user already exists
                        if User.objects.filter(email=row['email'].lower()).exists():
                            results['errors'].append({
                                'row': row_num,
                                'error': f"User with email {row['email']} already exists"
                            })
                            continue

                        # Create user
                        user = User.objects.create_user(
                            email=row['email'].lower(),
                            password=row.get('password', 'changeme123'),  # Default password
                            first_name=row.get('first_name', ''),
                            last_name=row.get('last_name', ''),
                            phone=row.get('phone', ''),
                            role=role
                        )

                        # Set optional fields
                        if row.get('date_of_birth'):
                            try:
                                user.date_of_birth = datetime.strptime(
                                    row['date_of_birth'],
                                    '%Y-%m-%d'
                                ).date()
                            except ValueError:
                                pass

                        if row.get('gender'):
                            user.gender = row['gender'].upper()

                        user.save()

                        results['success'].append({
                            'row': row_num,
                            'email': user.email,
                            'name': user.get_full_name()
                        })

                    except Exception as e:
                        results['errors'].append({
                            'row': row_num,
                            'error': str(e)
                        })

        except Exception as e:
            results['errors'].append({
                'row': 0,
                'error': f'File error: {str(e)}'
            })

        return results

    @staticmethod
    def import_users_excel(excel_file, role='STUDENT'):
        """
        Import users from Excel file
        """
        from .models import User

        results = {
            'success': [],
            'errors': [],
            'total': 0
        }

        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    results['total'] += 1

                    try:
                        # Convert row to dict
                        row_data = dict(zip(headers, row))

                        # Validate required fields
                        if not row_data.get('email'):
                            results['errors'].append({
                                'row': row_num,
                                'error': 'Email is required'
                            })
                            continue

                        # Check if user already exists
                        if User.objects.filter(email=row_data['email'].lower()).exists():
                            results['errors'].append({
                                'row': row_num,
                                'error': f"User with email {row_data['email']} already exists"
                            })
                            continue

                        # Create user
                        user = User.objects.create_user(
                            email=row_data['email'].lower(),
                            password=row_data.get('password', 'changeme123'),
                            first_name=row_data.get('first_name', ''),
                            last_name=row_data.get('last_name', ''),
                            phone=row_data.get('phone', ''),
                            role=role
                        )

                        # Set optional fields
                        if row_data.get('date_of_birth'):
                            if isinstance(row_data['date_of_birth'], datetime):
                                user.date_of_birth = row_data['date_of_birth'].date()
                            else:
                                try:
                                    user.date_of_birth = datetime.strptime(
                                        str(row_data['date_of_birth']),
                                        '%Y-%m-%d'
                                    ).date()
                                except ValueError:
                                    pass

                        if row_data.get('gender'):
                            user.gender = str(row_data['gender']).upper()

                        user.save()

                        results['success'].append({
                            'row': row_num,
                            'email': user.email,
                            'name': user.get_full_name()
                        })

                    except Exception as e:
                        results['errors'].append({
                            'row': row_num,
                            'error': str(e)
                        })

        except Exception as e:
            results['errors'].append({
                'row': 0,
                'error': f'File error: {str(e)}'
            })

        return results


class BulkExporter:
    """
    Handle bulk export operations
    """

    @staticmethod
    def export_users_csv(queryset):
        """
        Export users to CSV
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)

        # Write header
        writer.writerow([
            'Email', 'First Name', 'Last Name', 'Phone', 'Role',
            'Gender', 'Date of Birth', 'Date Joined', 'Is Active', 'Campus'
        ])

        # Write data
        for user in queryset:
            writer.writerow([
                user.email,
                user.first_name,
                user.last_name,
                user.phone or '',
                user.get_role_display(),
                user.gender or '',
                user.date_of_birth or '',
                user.date_joined.strftime('%Y-%m-%d %H:%M'),
                'Yes' if user.is_active else 'No',
                user.campus.name if user.campus else ''
            ])

        return response

    @staticmethod
    def export_users_excel(queryset):
        """
        Export users to Excel with formatting
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users Export"

        # Define header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = [
            'Email', 'First Name', 'Last Name', 'Phone', 'Role',
            'Gender', 'Date of Birth', 'Date Joined', 'Is Active', 'Campus'
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_num, user in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=user.email)
            ws.cell(row=row_num, column=2, value=user.first_name)
            ws.cell(row=row_num, column=3, value=user.last_name)
            ws.cell(row=row_num, column=4, value=user.phone or '')
            ws.cell(row=row_num, column=5, value=user.get_role_display())
            ws.cell(row=row_num, column=6, value=user.gender or '')
            ws.cell(row=row_num, column=7, value=user.date_of_birth or '')
            ws.cell(row=row_num, column=8, value=user.date_joined.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=9, value='Yes' if user.is_active else 'No')
            ws.cell(row=row_num, column=10, value=user.campus.name if user.campus else '')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        return response

    @staticmethod
    def get_import_template_csv():
        """
        Generate CSV template for user import
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'

        writer = csv.writer(response)

        # Write header
        writer.writerow([
            'email', 'first_name', 'last_name', 'phone',
            'date_of_birth', 'gender', 'password'
        ])

        # Write sample row
        writer.writerow([
            'john.doe@example.com', 'John', 'Doe', '+23276123456',
            '2000-01-15', 'MALE', 'changeme123'
        ])

        return response

    @staticmethod
    def get_import_template_excel():
        """
        Generate Excel template for user import
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Import Template"

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = [
            'email', 'first_name', 'last_name', 'phone',
            'date_of_birth', 'gender', 'password'
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write sample row
        ws.cell(row=2, column=1, value='john.doe@example.com')
        ws.cell(row=2, column=2, value='John')
        ws.cell(row=2, column=3, value='Doe')
        ws.cell(row=2, column=4, value='+23276123456')
        ws.cell(row=2, column=5, value='2000-01-15')
        ws.cell(row=2, column=6, value='MALE')
        ws.cell(row=2, column=7, value='changeme123')

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        ws_instructions.cell(row=1, column=1, value="Import Instructions:")
        ws_instructions.cell(row=2, column=1, value="1. Fill in user data starting from row 2")
        ws_instructions.cell(row=3, column=1, value="2. Required fields: email, first_name, last_name")
        ws_instructions.cell(row=4, column=1, value="3. Date format: YYYY-MM-DD")
        ws_instructions.cell(row=5, column=1, value="4. Gender: MALE, FEMALE, or OTHER")
        ws_instructions.cell(row=6, column=1, value="5. Default password is 'changeme123' if not specified")

        # Auto-adjust widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_import_template.xlsx"'

        wb.save(response)
        return response
